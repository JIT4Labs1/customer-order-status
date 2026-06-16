#!/usr/bin/env python3
"""
JIT4You Open Orders by Customer Report
=======================================
Fetches all 2026 non-cancelled Sales Orders from Vtiger CRM,
computes open (undelivered) items, groups by Customer > Vendor,
and generates an HTML email report + Excel attachment.

Sends via Zapier webhook to customersupport@jit4you.com.

Usage:
  python open_orders_report.py              # Normal run
  python open_orders_report.py --no-email   # Generate files only, don't send email
  python open_orders_report.py --dry-run    # Preview counts without generating report
"""

import json, base64, time, urllib.parse, urllib.request, ssl, os, sys, argparse
import concurrent.futures
from datetime import datetime, timezone
from collections import defaultdict

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
CONFIG = {
    # REST API (returns all SOs, unlike webservice.php which returns a limited set)
    "vtiger_rest_base":  "https://jit4youinc.od2.vtiger.com/restapi/v1/vtiger/default",
    # Secrets are read from environment variables (set as GitHub Actions secrets
    # when running in CI). Hardcoded values remain as a fallback for local Mac runs.
    "vtiger_user":       os.environ.get("VTIGER_USER", "customersupport@jit4you.com"),
    "vtiger_accesskey":  os.environ.get("VTIGER_ACCESS_KEY", ""),

    # ConMed vendor ID to exclude
    "conmed_vendor_id": "11x63346",

    # Purchase Orders to exclude from the report entirely (matched case-insensitively,
    # with or without the "PO" prefix). Excluded POs are skipped during PO matching,
    # so they never appear as pending POs, in vendor PO counts, or in vendor emails.
    "excluded_pos": ["PO394"],

    # Zapier webhook for sending the summary email
    "zapier_email_webhook": os.environ.get("ZAPIER_WEBHOOK_URL", ""),

    # Per-vendor "all open orders" HTML-table emails are sent here (via Zapier)
    # for internal review — NOT to the vendors. Set to "" to disable these sends.
    "vendor_drafts_to": "customersupport@jit4you.com",

    # BCC included on the per-PO "Email vendor" mailto: drafts (set to "" to omit).
    # The button opens a draft in your OS default mail reader. To route to the
    # Mac Outlook desktop app: System Settings → Desktop & Dock → Default email
    # reader → Microsoft Outlook (one-time setting).
    "vendor_followup_bcc": "",

    # Per-vendor contact first names used in the email greeting ("Hi Karen,").
    # Keys are matched case-insensitively as substrings against the vendor name
    # in Vtiger, so "aldx" matches "ALDX HOLDING CORPORATION".
    # If no key matches, the greeting falls back to "Hi <vendor name>,".
    "vendor_contact_first_names": {
        "aldx":      "Karen",
        "pma":       "Debbie",
        "allora":    "Arthur",
        "clearchem": "Reanna",
    },

    # Per-vendor "To:" address overrides. Keys match the vendor name the same way
    # as vendor_contact_first_names. When set, the override replaces whatever
    # email is on the Vtiger Vendor record.
    "vendor_contact_emails": {
        "pma": "debbie@pmaservices.com",
    },

    # Rate limiting. Vtiger 429s aggressively, so per-record retrieves are issued
    # sequentially with this delay. Results are cached to disk and resumed across
    # runs, so a conservative delay here is safe even though it's slower per pass.
    "delay_between_calls": 0.35,

    # Output directory (same folder as this script)
    "output_dir": os.path.dirname(os.path.abspath(__file__)),
}

VTIGER_BASE = "https://jit4youinc.od2.vtiger.com"
SKIP_ITEMS = ['shipping', 'tax', 'ca sales tax']

ctx = ssl.create_default_context()

# ─────────────────────────────────────────────
# UTILITY
# ─────────────────────────────────────────────
def http_request(url, method="GET", headers=None, data=None, json_body=None):
    """Make an HTTP request and return parsed JSON."""
    if headers is None:
        headers = {}
    if json_body is not None:
        data = json.dumps(json_body).encode("utf-8")
        headers["Content-Type"] = "application/json"
    elif data and isinstance(data, dict):
        data = urllib.parse.urlencode(data).encode("utf-8")
    elif data and isinstance(data, str):
        data = data.encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
            body = resp.read().decode("utf-8")
            return json.loads(body) if body.strip() else {}
    except urllib.error.HTTPError as e:
        error_body = e.read().decode("utf-8") if e.fp else ""
        log(f"  HTTP {e.code} error: {error_body[:300]}")
        raise

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


# ─────────────────────────────────────────────
# VTIGER REST API (Basic Auth — returns full dataset)
# ─────────────────────────────────────────────
class VtigerAPI:
    def __init__(self, rest_base, user, accesskey, cache_path=None):
        self.rest_base = rest_base.rstrip("/")
        self.user = user
        self.accesskey = accesskey
        # Basic Auth header: base64(user:accesskey)
        creds = base64.b64encode(f"{user}:{accesskey}".encode()).decode()
        self.auth_headers = {"Authorization": f"Basic {creds}"}
        # ── Persistent retrieve cache ──────────────────────────────
        # Vtiger rate-limits aggressively (429) and each automated run is
        # capped at ~45s, so a single pass can't fetch every record. We
        # cache per-record retrieve() results to disk keyed by record id.
        # Across repeated invocations the cache fills until a run can build
        # the full report from cache alone. Bulk queries are NOT cached
        # (they're fast and we want them fresh each run).
        self.cache_path = cache_path
        self.retrieve_cache = {}
        self.query_cache = {}
        self._uncached_fetches = 0
        self.fetch_failures = 0  # count of retrieves that failed this run (e.g. 429)
        if cache_path and os.path.exists(cache_path):
            try:
                with open(cache_path, "r") as f:
                    blob = json.load(f)
                self.retrieve_cache = blob.get("retrieve", {})
                self.query_cache = blob.get("query", {})
                log(f"  Loaded cache: {len(self.retrieve_cache)} records, "
                    f"{len(self.query_cache)} queries")
            except Exception as e:
                log(f"  Warning: could not load cache ({e}); starting fresh")
                self.retrieve_cache, self.query_cache = {}, {}

    def save_cache(self):
        if not self.cache_path:
            return
        try:
            tmp = self.cache_path + ".tmp"
            with open(tmp, "w") as f:
                json.dump({"retrieve": self.retrieve_cache,
                           "query": self.query_cache}, f)
            os.replace(tmp, self.cache_path)
        except Exception as e:
            log(f"  Warning: could not save cache ({e})")

    def login(self):
        """Verify REST API connectivity (no challenge/login needed for Basic Auth)."""
        # Simple test query to verify credentials work
        try:
            self.query("SELECT salesorder_no FROM SalesOrder LIMIT 0, 1;")
            log("Vtiger REST API: Connected successfully")
        except Exception as e:
            raise Exception(f"Vtiger REST API connection failed: {e}")

    def query(self, sql, max_retries=5):
        """Run a Vtiger SQL query with exponential-backoff retry on HTTP 429.
        Results are cached (date-scoped cache file) so repeated resume runs
        within the same day skip the slow paginated bulk queries."""
        if sql in self.query_cache:
            return self.query_cache[sql]
        url = f"{self.rest_base}/query?query={urllib.parse.quote(sql)}"
        last_err = None
        for attempt in range(max_retries):
            try:
                resp = http_request(url, headers=dict(self.auth_headers))
                if not resp.get("success"):
                    raise Exception(f"Vtiger query failed: {resp}")
                self.query_cache[sql] = resp["result"]
                self.save_cache()
                return resp["result"]
            except urllib.error.HTTPError as e:
                last_err = e
                if e.code == 429 and attempt < max_retries - 1:
                    wait = 2 ** attempt  # 1, 2, 4, 8, 16 seconds
                    log(f"  Query hit 429, retrying in {wait}s (attempt {attempt+1}/{max_retries})...")
                    time.sleep(wait)
                    continue
                raise
        raise last_err if last_err else Exception("query: exhausted retries")

    def query_all(self, sql_template, delay=0.15):
        """Paginate a query using LIMIT offset,100 until no more results."""
        all_results = []
        offset = 0
        while True:
            sql = f"{sql_template} LIMIT {offset}, 100;"
            results = self.query(sql)
            if not results:
                break
            all_results.extend(results)
            if len(results) < 100:
                break
            offset += 100
            time.sleep(delay)
        return all_results

    def retrieve(self, record_id):
        url = f"{self.rest_base}/retrieve?id={urllib.parse.quote(record_id)}"
        resp = http_request(url, headers=dict(self.auth_headers))
        if not resp.get("success"):
            raise Exception(f"Vtiger retrieve failed for {record_id}: {resp}")
        return resp["result"]

    def retrieve_with_retry(self, record_id, label="record", max_retries=6):
        """Retrieve one record, using the persistent cache when available and
        retrying on HTTP 429 with exponential backoff. Returns the record dict,
        or None on failure. Successful fetches are cached and flushed to disk
        periodically so progress survives across capped invocations."""
        if record_id in self.retrieve_cache:
            return self.retrieve_cache[record_id]
        for attempt in range(max_retries):
            try:
                detail = self.retrieve(record_id)
                self.retrieve_cache[record_id] = detail
                self._uncached_fetches += 1
                if self._uncached_fetches % 3 == 0:
                    self.save_cache()
                return detail
            except urllib.error.HTTPError as e:
                if e.code == 429 and attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
                    continue
                log(f"  Warning: Failed to retrieve {label} {record_id}: HTTP {e.code}")
                self.fetch_failures += 1
                return None
            except Exception as e:
                log(f"  Warning: Failed to retrieve {label} {record_id}: {e}")
                self.fetch_failures += 1
                return None
        self.fetch_failures += 1
        return None

    def retrieve_many(self, record_ids, label="record", deadline=None):
        """Retrieve many records sequentially (gentle on Vtiger's rate limiter),
        serving cache hits instantly. Returns {id: detail|None}. Records not
        fetched before `deadline` (epoch seconds) are left out of the result so
        the caller can finish the current run and resume next time from cache."""
        ids = list(record_ids)
        results = {}
        delay = CONFIG.get("delay_between_calls", 0.3)
        for rid in ids:
            cached = rid in self.retrieve_cache
            if deadline and not cached and time.time() > deadline:
                # Out of time for fresh fetches this run; stop early.
                break
            results[rid] = self.retrieve_with_retry(rid, label)
            if not cached:
                time.sleep(delay)
        self.save_cache()
        return results


# ─────────────────────────────────────────────
# DATA EXTRACTION
# ─────────────────────────────────────────────
def extract_open_orders(vt, dry_run=False):
    """Full fresh extraction of open orders from Vtiger."""

    # STEP 1: Get all Sales Orders (lightweight query first, then filter)
    # NOTE: SELECT * hits API response size limits and returns truncated results.
    # Instead, query only the fields needed for filtering, then retrieve full details later.
    log("Step 1: Fetching all Sales Orders...")
    all_sos_raw = vt.query_all(
        "SELECT id, salesorder_no, subject, sostatus, createdtime, modifiedtime, duedate, account_id "
        "FROM SalesOrder"
    )

    # Filter to 2026 and non-cancelled locally (based on creation date only)
    def is_2026(s):
        created = s.get("createdtime", "")
        return "2026" in str(created)

    all_sos = [s for s in all_sos_raw if is_2026(s)]
    non_cancelled = [s for s in all_sos if s.get("sostatus") != "Cancelled"]
    log(f"  Found {len(all_sos_raw)} total SOs, {len(all_sos)} in 2026, {len(non_cancelled)} non-cancelled")

    if dry_run:
        log(f"  SO numbers: {[s.get('salesorder_no', s.get('subject','?')) for s in non_cancelled]}")

    # STEP 2: ConMed customer exclusion
    log("Step 2: Checking ConMed customer exclusion...")
    unique_accounts = set(s.get("account_id", "") for s in non_cancelled if s.get("account_id"))
    conmed_accounts = set()
    # Get all POs with ConMed as vendor in one query (lightweight fields)
    try:
        conmed_pos = vt.query_all(
            f"SELECT id, postatus, contact_id FROM PurchaseOrder "
            f"WHERE vendor_id = '{CONFIG['conmed_vendor_id']}'"
        )
        # Filter out cancelled POs locally, collect account IDs
        for po in conmed_pos:
            if po.get("postatus", "") != "Cancelled":
                cid = po.get("contact_id", "")
                if cid:
                    conmed_accounts.add(cid)
    except Exception as e:
        log(f"  Warning: ConMed PO query failed: {e}")
        # Fallback: check per account
        for acct_id in unique_accounts:
            try:
                pos = vt.query(
                    f"SELECT id, postatus FROM PurchaseOrder WHERE vendor_id = '{CONFIG['conmed_vendor_id']}' "
                    f"AND contact_id = '{acct_id}' LIMIT 0, 100;"
                )
                active_pos = [p for p in pos if p.get("postatus", "") != "Cancelled"]
                if active_pos:
                    conmed_accounts.add(acct_id)
            except Exception:
                pass
            time.sleep(CONFIG["delay_between_calls"])

    included_sos = [s for s in non_cancelled if s.get("account_id") not in conmed_accounts]
    log(f"  Excluded {len(conmed_accounts)} ConMed customers, {len(included_sos)} SOs remain")

    if dry_run:
        log("DRY RUN - stopping here")
        return []

    # STEP 3: Resolve customer names (bulk query — one API call instead of N retrieves)
    log("Step 3: Resolving customer names...")
    acct_ids = set(s.get("account_id", "") for s in included_sos if s.get("account_id"))
    acct_names = {}
    try:
        all_accounts = vt.query_all("SELECT id, accountname FROM Accounts")
        acct_names = {a["id"]: a.get("accountname", "Unknown")
                      for a in all_accounts if a.get("id") in acct_ids}
    except Exception as e:
        log(f"  Warning: Bulk account query failed ({e}), falling back to individual lookups")
        for acct_id in acct_ids:
            try:
                acct = vt.retrieve(acct_id)
                acct_names[acct_id] = acct.get("accountname", "Unknown")
            except Exception:
                acct_names[acct_id] = "Unknown"
            time.sleep(CONFIG["delay_between_calls"])
    for aid in acct_ids:
        acct_names.setdefault(aid, "Unknown")
    log(f"  Resolved {len(acct_names)} customer names")

    # STEP 4: Retrieve SO details with line items
    log("Step 4: Retrieving SO details + line items...")
    so_details = {}
    all_product_ids = set()
    so_detail_by_id = vt.retrieve_many([so["id"] for so in included_sos], label="SO")
    for so in included_sos:
        detail = so_detail_by_id.get(so["id"])
        if detail:
            so_num = detail.get("salesorder_no", so.get("salesorder_no", so.get("subject", "")))
            so_details[so_num] = detail
            line_items = detail.get("LineItems", detail.get("lineItems", []))
            if isinstance(line_items, list):
                for li in line_items:
                    pid = li.get("productid", "")
                    if pid:
                        all_product_ids.add(pid)
    log(f"  Retrieved {len(so_details)} SO details, {len(all_product_ids)} unique products")

    # STEP 5: Resolve product info (bulk query — one paginated call instead of N retrieves)
    log("Step 5: Resolving product names and vendor IDs...")
    product_info = {}  # pid -> {name, vendor_id}
    vendor_ids = set()
    try:
        all_products = vt.query_all("SELECT id, productname, vendor_id FROM Products")
        for p in all_products:
            pid = p.get("id", "")
            if pid in all_product_ids:
                vid = p.get("vendor_id", "")
                product_info[pid] = {
                    "name": p.get("productname", p.get("label", "")),
                    "vendor_id": vid,
                }
                if vid and vid != "0":
                    vendor_ids.add(vid)
    except Exception as e:
        log(f"  Warning: Bulk product query failed ({e}), falling back to individual lookups")
        for pid in all_product_ids:
            try:
                prod = vt.retrieve(pid)
                vid = prod.get("vendor_id", "")
                product_info[pid] = {
                    "name": prod.get("productname", prod.get("label", "")),
                    "vendor_id": vid,
                }
                if vid and vid != "0":
                    vendor_ids.add(vid)
            except Exception:
                product_info[pid] = {"name": "", "vendor_id": ""}
            time.sleep(CONFIG["delay_between_calls"])
    for pid in all_product_ids:
        product_info.setdefault(pid, {"name": "", "vendor_id": ""})
    log(f"  Resolved {len(product_info)} products, {len(vendor_ids)} unique vendors")

    # STEP 6: Resolve vendor names + emails (bulk query — one paginated call instead of N retrieves)
    log("Step 6: Resolving vendor names + emails...")
    vendor_names = {}
    vendor_emails = {}  # vid -> email
    try:
        all_vendors = vt.query_all("SELECT id, vendorname, email FROM Vendors")
        for v in all_vendors:
            if v.get("id"):
                vendor_names[v["id"]] = v.get("vendorname", "Unknown")
                vendor_emails[v["id"]] = v.get("email", "") or ""
    except Exception as e:
        log(f"  Warning: Bulk vendor query failed ({e}), falling back to individual lookups")
        for vid in vendor_ids:
            try:
                vendor = vt.retrieve(vid)
                vendor_names[vid] = vendor.get("vendorname", vendor.get("label", "Unknown"))
                vendor_emails[vid] = vendor.get("email", "") or ""
            except Exception:
                vendor_names[vid] = "Unknown"
                vendor_emails[vid] = ""
            time.sleep(CONFIG["delay_between_calls"])
    _emailed = sum(1 for e in vendor_emails.values() if e)
    log(f"  Vendor names resolved: {len(vendor_names)} vendors ({_emailed} have email on file)")

    # Build product -> vendor name map and vendor name -> email map
    product_vendor = {}
    for pid, info in product_info.items():
        product_vendor[pid] = vendor_names.get(info["vendor_id"], "Unspecified")

    vendor_name_to_email = {}
    for vid, name in vendor_names.items():
        em = vendor_emails.get(vid, "")
        if em and (name not in vendor_name_to_email or not vendor_name_to_email[name]):
            vendor_name_to_email[name] = em

    # Apply per-vendor email overrides from CONFIG (e.g. PMA -> debbie@pmaservices.com).
    # Pattern keys match case-insensitively as substrings against the vendor name.
    email_overrides = CONFIG.get("vendor_contact_emails", {})
    if email_overrides:
        for name in list(vendor_name_to_email.keys()):
            name_lower = name.lower()
            for key, override in email_overrides.items():
                if key.lower() in name_lower:
                    vendor_name_to_email[name] = override
                    log(f"  Vendor email override: {name} -> {override}")
                    break
        # Also seed entries for vendors that had no email on file but match an override
        for name in {n for n in vendor_names.values()}:
            name_lower = name.lower()
            if not vendor_name_to_email.get(name):
                for key, override in email_overrides.items():
                    if key.lower() in name_lower:
                        vendor_name_to_email[name] = override
                        log(f"  Vendor email seeded: {name} -> {override}")
                        break

    # STEP 7: Get delivered Delivery Notes (lightweight query, filter locally)
    log("Step 7: Fetching Delivery Notes...")
    all_dns_raw = vt.query_all(
        "SELECT id, deliverynote_status, related_to, createdtime, modifiedtime "
        "FROM DeliveryNotes"
    )

    # Filter to 2026 and Delivered status locally
    all_dns = [dn for dn in all_dns_raw
               if any("2026" in str(dn.get(f, "")) for f in ['createdtime', 'modifiedtime'])
               and dn.get("deliverynote_status", "") == "Delivered"]
    log(f"  Found {len(all_dns_raw)} total DNs, {len(all_dns)} delivered in 2026")

    # Filter to DNs linked to our SOs
    so_id_set = set(s["id"] for s in included_sos)
    so_id_to_num = {s["id"]: s.get("salesorder_no", s.get("subject", "")) for s in included_sos}
    relevant_dns = [dn for dn in all_dns if dn.get("related_to") in so_id_set]
    log(f"  {len(relevant_dns)} linked to our SOs")

    # Retrieve DN details for line items
    delivered_map = defaultdict(lambda: defaultdict(float))  # so_num -> product_id -> qty
    dn_detail_by_id = vt.retrieve_many([dn["id"] for dn in relevant_dns], label="DN")
    for dn in relevant_dns:
        detail = dn_detail_by_id.get(dn["id"])
        if detail:
            so_id = detail.get("related_to", dn.get("related_to", ""))
            so_num = so_id_to_num.get(so_id, "")
            if so_num:
                line_items = detail.get("LineItems", detail.get("lineItems", []))
                if isinstance(line_items, list):
                    for li in line_items:
                        pid = li.get("productid", "")
                        qty = float(li.get("quantity", li.get("qty", 0)))
                        if pid and qty > 0:
                            delivered_map[so_num][pid] += qty
    log(f"  Built delivery map for {len(delivered_map)} SOs")

    # STEP 8: Find Purchase Orders directly linked to our SO line items
    # Build SO record ID -> SO number map for matching
    log("Step 8: Matching Purchase Orders to SO line items...")
    so_id_to_num = {}  # "4x12345" -> "SO202"
    so_product_ids = set()  # all product IDs across our SOs
    for so_num, detail in so_details.items():
        so_id_to_num[detail.get("id", "")] = so_num
        line_items = detail.get("LineItems", detail.get("lineItems", []))
        if isinstance(line_items, list):
            for li in line_items:
                pid = li.get("productid", "")
                if pid:
                    so_product_ids.add(pid)

    # Try multiple possible field names for the SO link on POs
    # User confirmed POs have an "SO ID" field that holds the Sales Order ID
    so_link_fields = ["so_id", "soid", "salesorder_id", "cf_potentialrelso", "cf_salesorderid"]
    field_list = ", ".join(so_link_fields)

    # First try with all candidate fields; if any cause a 400, fall back
    po_query_fields = "id, purchaseorder_no, postatus, createdtime"
    for field in so_link_fields:
        try:
            vt.query(f"SELECT {field} FROM PurchaseOrder LIMIT 0, 1;")
            po_query_fields += f", {field}"
            log(f"  PO field '{field}' is accessible")
        except Exception:
            pass

    all_pos_raw = vt.query_all(f"SELECT {po_query_fields} FROM PurchaseOrder")

    # Filter: non-cancelled only
    active_pos = [p for p in all_pos_raw if p.get("postatus", "") != "Cancelled"]
    log(f"  Found {len(all_pos_raw)} total POs, {len(active_pos)} non-cancelled")

    # Match POs to SOs using any SO reference field
    so_product_to_pos = defaultdict(list)  # key: (so_num, product_id) -> [po_num]
    po_vendor_override = {}  # key: (so_num, product_id) -> vendor_name from PO
    po_created_date = {}  # key: (so_num, product_id) -> earliest PO createdtime string
    po_eta = {}  # key: (so_num, product_id) -> ETA date string from PO line item
    po_all_items = defaultdict(list)  # key: po_num -> [(pid, qty, linked_so), ...]
    matched_count = 0
    retrieved_count = 0

    # Determine which POs are linked to our SOs (local filter — no network)
    linked_pos = []  # [(po, linked_so), ...]
    for po in active_pos:
        linked_so = ""
        for field in so_link_fields:
            so_ref = po.get(field, "")
            if so_ref and so_ref in so_id_to_num:
                linked_so = so_id_to_num[so_ref]
                break
        if linked_so:
            linked_pos.append((po, linked_so))

    # Prefetch all linked PO details in parallel
    po_detail_by_id = vt.retrieve_many([po["id"] for po, _ in linked_pos], label="PO")

    for po, linked_so in linked_pos:
        detail = po_detail_by_id.get(po["id"])
        if detail is None:
            continue

        try:
            retrieved_count += 1
            po_num = detail.get("purchaseorder_no", po.get("purchaseorder_no", ""))

            if detail.get("postatus", "") == "Cancelled":
                continue

            # Skip explicitly excluded POs (e.g. PO394) — matched case-insensitively,
            # with or without the "PO" prefix, so they never appear in the report.
            _excl = {str(p).strip().upper().lstrip("PO").lstrip() for p in CONFIG.get("excluded_pos", [])}
            _excl_full = {str(p).strip().upper() for p in CONFIG.get("excluded_pos", [])}
            _pn_norm = str(po_num).strip().upper()
            if _pn_norm in _excl_full or _pn_norm.lstrip("PO").lstrip() in _excl:
                continue

            # Get the PO's vendor (most up-to-date vendor for these items)
            po_vendor_id = detail.get("vendor_id", "")
            po_vendor_name = ""
            if po_vendor_id:
                if po_vendor_id not in vendor_names:
                    try:
                        v = vt.retrieve(po_vendor_id)
                        vendor_names[po_vendor_id] = v.get("vendorname", v.get("label", "Unknown"))
                    except Exception:
                        vendor_names[po_vendor_id] = "Unknown"
                    time.sleep(CONFIG["delay_between_calls"])
                po_vendor_name = vendor_names.get(po_vendor_id, "")

            po_created = detail.get("createdtime", po.get("createdtime", ""))

            line_items = detail.get("LineItems", detail.get("lineItems", []))
            if isinstance(line_items, list):
                for li in line_items:
                    pid = li.get("productid", "")
                    if pid:
                        so_product_to_pos[(linked_so, pid)].append(po_num)
                        # Store the PO's vendor as the override for this SO+product
                        if po_vendor_name:
                            po_vendor_override[(linked_so, pid)] = po_vendor_name
                        # Store earliest PO created date
                        if po_created:
                            existing = po_created_date.get((linked_so, pid), "")
                            if not existing or po_created < existing:
                                po_created_date[(linked_so, pid)] = po_created
                        # Capture ETA from PO line item
                        eta_val = li.get("cf_purchaseorder_eta", "")
                        if eta_val and eta_val not in ("0000-00-00", ""):
                            existing_eta = po_eta.get((linked_so, pid), "")
                            # Keep the earliest ETA across multiple POs
                            if not existing_eta or eta_val < existing_eta:
                                po_eta[(linked_so, pid)] = eta_val
                        # Collect PO line items for PO-level delivery tracking
                        po_ord = float(li.get("quantity", li.get("qty", 0)))
                        po_all_items[po_num].append((pid, po_ord, linked_so))
                        matched_count += 1
        except Exception as e:
            log(f"  Warning: Error processing PO {po.get('purchaseorder_no', po.get('id', '?'))}: {e}")
    log(f"  Retrieved {retrieved_count} linked POs, matched {matched_count} line items to {len(so_product_to_pos)} (SO, product) pairs")
    log(f"  Vendor overrides from POs: {len(po_vendor_override)}")

    # Compute PO-level delivery status using delivered_map from Delivery Notes
    # For each PO, count how many of its line items are fully delivered
    po_delivery_status = {}  # po_num -> "delivered/total"
    for po_num, items in po_all_items.items():
        total = len(items)
        delivered_count = 0
        for pid, qty, so_num_ref in items:
            dn_qty = delivered_map.get(so_num_ref, {}).get(pid, 0)
            if dn_qty >= qty:
                delivered_count += 1
        po_delivery_status[po_num] = f"{delivered_count}/{total}"
    log(f"  PO delivery status computed for {len(po_delivery_status)} POs")

    # STEP 9: Compute open items
    log("Step 9: Computing open items...")
    open_items = []
    for so_num, detail in so_details.items():
        account_id = detail.get("account_id", "")
        customer = acct_names.get(account_id, "Unknown")
        so_status = detail.get("sostatus", "")
        created = detail.get("createdtime", "")
        order_date = created.split(" ")[0] if created else ""
        so_id_full = detail.get("id", "")
        numeric_id = so_id_full.split("x")[-1] if "x" in so_id_full else so_id_full

        line_items = detail.get("LineItems", detail.get("lineItems", []))
        if not isinstance(line_items, list):
            continue

        for li in line_items:
            pid = li.get("productid", "")
            if not pid:
                continue

            # Get product name
            pinfo = product_info.get(pid, {})
            product_name = pinfo.get("name", "") or li.get("productid_display", "") or li.get("productName", "")
            if not product_name:
                continue

            # Skip non-product items
            pname_lower = product_name.lower()
            if any(skip in pname_lower for skip in SKIP_ITEMS):
                continue
            if product_name.strip().replace(".", "").replace("%", "").isdigit():
                continue
            import re
            if re.match(r'^\d+(\.\d+)?%$', product_name.strip()):
                continue

            # Use PO vendor if available (most current), otherwise product vendor
            vendor_name = po_vendor_override.get((so_num, pid), "") or product_vendor.get(pid, "Unspecified")

            # Skip ConMed vendor products
            if "conmed" in vendor_name.lower():
                continue

            ordered_qty = float(li.get("quantity", li.get("qty", 0)))
            delivered_qty = delivered_map.get(so_num, {}).get(pid, 0)
            open_qty = ordered_qty - delivered_qty

            if open_qty <= 0:
                continue

            unit_price = float(li.get("listprice", li.get("price", 0)))
            open_value = open_qty * unit_price
            # Pending POs: only show POs directly linked to this SO + product
            linked_pos = so_product_to_pos.get((so_num, pid), [])
            pending_pos = ", ".join(sorted(set(linked_pos)))

            # ETA from PO line item
            eta = po_eta.get((so_num, pid), "")

            open_items.append({
                "customer": customer,
                "so_num": so_num,
                "so_status": so_status,
                "order_date": order_date,
                "so_id": numeric_id,
                "product": product_name,
                "vendor": vendor_name,
                "vendor_email": vendor_name_to_email.get(vendor_name, ""),
                "ordered_qty": ordered_qty,
                "delivered_qty": delivered_qty,
                "open_qty": open_qty,
                "unit_price": unit_price,
                "open_value": open_value,
                "pending_pos": pending_pos,
                "eta": eta,
            })

    # Sort by date ascending
    open_items.sort(key=lambda r: r["order_date"])
    log(f"  {len(open_items)} open items across {len(set(i['customer'] for i in open_items))} customers")
    return open_items


# ─────────────────────────────────────────────
# REPORT GENERATION
# ─────────────────────────────────────────────
def _fmt_eta_long(eta):
    """Format an ETA value (e.g. '2026-04-28 12:34:56') as 'Apr 28, 2026'. Falls back to 'NO ETA'/raw."""
    if not eta:
        return "NO ETA"
    s = str(eta).split(" ")[0]
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%b %d, %Y")
    except Exception:
        return s


def build_po_email_url(po_num, info, bcc_email=""):
    """Build a `mailto:` URL that opens a draft to the vendor with a plain-text
    bullet list of open items.

    The button delegates to the OS default mail reader. To make this open in
    the Mac Outlook desktop app, set Outlook as the default email reader:
        System Settings → Desktop & Dock → Default email reader → Microsoft Outlook
    Once that is set, the desktop Outlook (already signed in) opens directly.

    Returns "" if the vendor has no email on file.
    """
    vendor_email = info.get("vendor_email", "")
    if not vendor_email:
        return ""

    vendor = info.get("vendor", "team")
    items = info.get("items", [])

    # Resolve greeting: prefer mapped contact first name; otherwise use vendor name.
    greeting_name = vendor
    contacts = CONFIG.get("vendor_contact_first_names", {})
    if contacts and vendor:
        vendor_lower = vendor.lower()
        for key, first_name in contacts.items():
            if key.lower() in vendor_lower:
                greeting_name = first_name
                break

    body_lines = [
        f"Hi {greeting_name},",
        "",
        f"Following up on PO {po_num}. Could you please share an update on the open items below?",
        "",
    ]
    for it in items:
        qty = it.get("open_qty", 0)
        try:
            qty_str = str(int(qty)) if float(qty) == int(qty) else f"{qty:g}"
        except Exception:
            qty_str = str(qty)
        body_lines.append(
            f"- {it.get('product', '')} — {qty_str} open, ETA: {_fmt_eta_long(it.get('eta', ''))}"
        )
    body_lines.extend(["", "Thank you,", "Amir"])
    body = "\n".join(body_lines)
    subject = f"{po_num} ETA?"

    params = [("subject", subject), ("body", body)]
    if bcc_email:
        params.append(("bcc", bcc_email))
    query = urllib.parse.urlencode(params, quote_via=urllib.parse.quote, safe="")
    addr = urllib.parse.quote(vendor_email, safe="@.+-_")
    return f"mailto:{addr}?{query}"


def build_vendor_email_url(vendor, items, vendor_email):
    """Build a `mailto:` URL that opens a single draft to a vendor listing ALL
    of that vendor's open orders across every customer and PO.

    Uses only data already present in the report (no extra Vtiger calls). The
    button it powers sits next to each vendor in the top summary table; the user
    sends the draft manually. Returns "" if no vendor email is available.
    """
    if not vendor_email:
        return ""

    # Resolve greeting: prefer mapped contact first name; otherwise use vendor name.
    greeting_name = vendor
    contacts = CONFIG.get("vendor_contact_first_names", {})
    if contacts and vendor:
        vendor_lower = vendor.lower()
        for key, first_name in contacts.items():
            if key.lower() in vendor_lower:
                greeting_name = first_name
                break

    def _qstr(qty):
        try:
            return str(int(qty)) if float(qty) == int(qty) else f"{qty:g}"
        except Exception:
            return str(qty)

    # Group this vendor's open items by PO number for a readable draft.
    by_po = defaultdict(list)
    no_po = []
    for it in items:
        pos = [p.strip() for p in (it.get("pending_pos", "") or "").split(",") if p.strip()]
        if pos:
            for p in pos:
                by_po[p].append(it)
        else:
            no_po.append(it)

    body_lines = [
        f"Hi {greeting_name},",
        "",
        "Following up on all of our open orders with you below. Could you please "
        "share an ETA update on each of the open items?",
        "",
    ]

    def _emit_item(it):
        eta = _fmt_eta_long(it.get("eta", "")) if it.get("eta", "") else "NO ETA"
        body_lines.append(
            f"  - [{it.get('order_date', '')}] {it.get('product', '')} ({it.get('customer', '')}) — "
            f"Ordered {_qstr(it.get('ordered_qty', 0))}, "
            f"Delivered {_qstr(it.get('delivered_qty', 0))}, "
            f"Open {_qstr(it.get('open_qty', 0))}, ETA: {eta}"
        )

    def _min_date(its):
        ds = [it.get("order_date", "") for it in its if it.get("order_date")]
        return min(ds) if ds else "9999-99-99"
    for its in by_po.values():
        its.sort(key=lambda it: (it.get("order_date", ""), it.get("product", "")))
    for po in sorted(by_po.keys(), key=lambda p: _min_date(by_po[p])):
        body_lines.append(f"{po}:")
        for it in by_po[po]:
            _emit_item(it)
        body_lines.append("")
    if no_po:
        no_po.sort(key=lambda it: (it.get("order_date", ""), it.get("product", "")))
        body_lines.append("Pending PO (not yet assigned):")
        for it in no_po:
            _emit_item(it)
        body_lines.append("")

    body_lines.extend(["Thank you,", "Amir"])
    body = "\n".join(body_lines)
    subject = f"JIT4You open orders — ETA update request"

    params = [("subject", subject), ("body", body)]
    bcc_email = CONFIG.get("vendor_followup_bcc", "")
    if bcc_email:
        params.append(("bcc", bcc_email))
    query = urllib.parse.urlencode(params, quote_via=urllib.parse.quote, safe="")
    addr = urllib.parse.quote(vendor_email, safe="@.+-_")
    return f"mailto:{addr}?{query}"


def _vendor_greeting(vendor):
    """Resolve the greeting first name for a vendor (mapped name, else vendor name)."""
    greeting_name = vendor
    contacts = CONFIG.get("vendor_contact_first_names", {})
    if contacts and vendor:
        vendor_lower = vendor.lower()
        for key, first_name in contacts.items():
            if key.lower() in vendor_lower:
                greeting_name = first_name
                break
    return greeting_name


def build_vendor_email_html(vendor, items):
    """Build a polished HTML-table email body listing ALL of a vendor's open
    orders across every customer/PO. Used for the per-vendor Gmail drafts the
    user reviews and sends manually. Uses only data already in the report.
    """
    greeting_name = _vendor_greeting(vendor)

    def _qstr(qty):
        try:
            return str(int(qty)) if float(qty) == int(qty) else f"{qty:g}"
        except Exception:
            return str(qty)

    # Group this vendor's items by PO number; items with no PO go in a final group.
    by_po = defaultdict(list)
    no_po = []
    for it in items:
        pos = [p.strip() for p in (it.get("pending_pos", "") or "").split(",") if p.strip()]
        if pos:
            for p in pos:
                by_po[p].append(it)
        else:
            no_po.append(it)

    def _item_row(it):
        td = "padding:7px 10px;border:1px solid #d8dee4;"
        return (
            f'<tr>'
            f'<td style="{td}white-space:nowrap;">{it.get("order_date","")}</td>'
            f'<td style="{td}">{it.get("product","")}</td>'
            f'<td style="{td}">{it.get("customer","")}</td>'
            f'<td style="{td}text-align:center;">{_qstr(it.get("ordered_qty",0))}</td>'
            f'<td style="{td}text-align:center;">{_qstr(it.get("delivered_qty",0))}</td>'
            f'<td style="{td}text-align:center;font-weight:700;color:#c0392b;">{_qstr(it.get("open_qty",0))}</td>'
            f'</tr>'
        )

    NCOL = 6  # Order Date, Product, Customer, Ordered, Delivered, Open
    po_hdr = ('padding:7px 10px;border:1px solid #1F4E79;background:#e8eef4;'
              'color:#1F4E79;font-weight:700;font-size:12px;')
    rows_html = ""
    # Sort items within each PO group by order date ascending, and order the PO
    # groups themselves by their earliest order date ascending.
    def _min_date(its):
        ds = [it.get("order_date", "") for it in its if it.get("order_date")]
        return min(ds) if ds else "9999-99-99"
    for its in by_po.values():
        its.sort(key=lambda it: (it.get("order_date", ""), it.get("product", "")))
    groups = sorted(((po, by_po[po]) for po in by_po.keys()), key=lambda g: _min_date(g[1]))
    if no_po:
        no_po.sort(key=lambda it: (it.get("order_date", ""), it.get("product", "")))
        groups.append(("No PO assigned", no_po))
    for po, its in groups:
        rows_html += f'<tr><td colspan="{NCOL}" style="{po_hdr}">PO {po}</td></tr>' if po != "No PO assigned" \
            else f'<tr><td colspan="{NCOL}" style="{po_hdr}">{po}</td></tr>'
        for it in its:
            rows_html += _item_row(it)

    th = ('padding:8px 10px;border:1px solid #0D2B45;color:#fff;text-align:left;'
          'font-size:12px;font-weight:700;')
    html = (
        f'<div style="font-family:Arial,Helvetica,sans-serif;font-size:13px;color:#2c3e50;line-height:1.5;">'
        f'<p>Hi {greeting_name},</p>'
        f'<p>Following up on all of our open orders with you. Could you please share an '
        f'ETA update on the open items below?</p>'
        f'<table cellspacing="0" cellpadding="0" style="border-collapse:collapse;width:100%;'
        f'max-width:760px;font-size:12px;font-family:Arial,Helvetica,sans-serif;background:#ffffff;">'
        f'<thead><tr style="background:#0D2B45;">'
        f'<th style="{th}">Order Date</th>'
        f'<th style="{th}">Product</th>'
        f'<th style="{th}">Customer</th><th style="{th}text-align:center;">Ordered</th>'
        f'<th style="{th}text-align:center;">Delivered</th><th style="{th}text-align:center;">Open</th>'
        f'</tr></thead><tbody>{rows_html}</tbody></table>'
        f'<p>Thank you,<br>Amir</p></div>'
    )
    return html


def build_inpage_sender_html(vendor_drafts, deliver_to, webhook_url):
    """Build the self-contained sender embedded in the report (hidden form + iframe
    + a small fixed status banner + ooSend()). Each per-vendor button calls
    ooSend(index); clicking it (in a browser, e.g. the report opened as a local
    file) POSTs that vendor's styled HTML email to the Zapier webhook for delivery
    to `deliver_to`. A form POST is used (not fetch) so it works from file:// too.
    """
    arr = [
        {"vendor": d["vendor"], "intended_to": d["to"],
         "html_body": d["html_body"], "item_count": d["item_count"]}
        for d in vendor_drafts
    ]
    payload_json = json.dumps(arr).replace("</", "<\\/").replace("<!--", "<\\!--")
    deliver_to_js = json.dumps(deliver_to)

    return (
        "<div id=\"oo-status\" style=\"display:none;position:fixed;left:50%;bottom:24px;"
        "transform:translateX(-50%);z-index:9999;max-width:90%;padding:12px 20px;border-radius:8px;"
        "font-family:Arial,Helvetica,sans-serif;font-size:14px;font-weight:700;box-shadow:0 4px 18px rgba(0,0,0,.2);\"></div>"
        "<iframe name=\"oo_sink\" style=\"display:none;\"></iframe>"
        f"<form id=\"oo_form\" method=\"POST\" action=\"{webhook_url}\" "
        "target=\"oo_sink\" accept-charset=\"UTF-8\" style=\"display:none;\">"
        "<input type=\"hidden\" name=\"to\" id=\"oo_to\">"
        "<input type=\"hidden\" name=\"subject\" id=\"oo_subject\">"
        "<input type=\"hidden\" name=\"html_body\" id=\"oo_html\">"
        "</form>"
        "<script>"
        f"var OO_VENDORS={payload_json};"
        f"var OO_DELIVER_TO={deliver_to_js};"
        "function ooBanner(msg,kind){var s=document.getElementById('oo-status');"
        "var c={info:['#fff3cd','#856404','#ffeeba'],ok:['#d4edda','#155724','#c3e6cb'],"
        "err:['#f8d7da','#721c24','#f5c6cb']}[kind||'info'];"
        "s.style.background=c[0];s.style.color=c[1];s.style.border='1px solid '+c[2];"
        "s.style.display='block';s.textContent=msg;}"
        "function ooSend(i){var d=OO_VENDORS[i];if(!d){ooBanner('Vendor not found.','err');return;}"
        "ooBanner('Sending '+d.vendor+' email to '+OO_DELIVER_TO+'…','info');"
        "document.getElementById('oo_to').value=OO_DELIVER_TO;"
        "document.getElementById('oo_subject').value='[Open-orders draft \\u2192 '+d.vendor+' | '+d.intended_to+'] ETA update request';"
        "document.getElementById('oo_html').value=d.html_body;"
        "try{document.getElementById('oo_form').submit();}catch(e){ooBanner('Send failed: '+e,'err');return;}"
        "setTimeout(function(){ooBanner('\\u2713 Sent '+d.vendor+' ('+d.item_count+' items) to '+OO_DELIVER_TO+'. Check your inbox.','ok');},1200);}"
        "</script>"
    )


def generate_report(open_items, output_dir):
    """Generate HTML email + Excel report with grouped layout."""

    # Group: Customer -> Vendor -> items
    by_customer = defaultdict(lambda: defaultdict(list))
    for item in open_items:
        by_customer[item["customer"]][item["vendor"]].append(item)

    total_customers = len(by_customer)
    total_sos = len(set((i["customer"], i["so_num"]) for i in open_items))
    total_items = len(open_items)
    report_date = datetime.now().strftime("%B %d, %Y")

    # Compute open PO count per vendor (unique PO numbers per vendor)
    vendor_po_count = defaultdict(set)
    for item in open_items:
        if item["pending_pos"]:
            for po_num in item["pending_pos"].split(", "):
                po_num = po_num.strip()
                if po_num:
                    vendor_po_count[item["vendor"]].add(po_num)

    # Build PO -> {vendor, vendor_email, items[], customers[]} for the per-PO email buttons
    po_email_data = {}
    for item in open_items:
        if not item.get("pending_pos"):
            continue
        for po_num in item["pending_pos"].split(", "):
            po_num = po_num.strip()
            if not po_num:
                continue
            entry = po_email_data.setdefault(po_num, {
                "vendor": item["vendor"],
                "vendor_email": item.get("vendor_email", ""),
                "customers": [],
                "items": [],
            })
            if item["customer"] not in entry["customers"]:
                entry["customers"].append(item["customer"])
            entry["items"].append({
                "product": item["product"],
                "open_qty": item["open_qty"],
                "eta": item.get("eta", ""),
                "so_num": item["so_num"],
                "customer": item["customer"],
            })
    vendor_po_summary = sorted(
        [(v, len(pos)) for v, pos in vendor_po_count.items()],
        key=lambda x: x[1], reverse=True
    )

    # Per-vendor aggregation for the top-table "Email all open orders" button:
    # every open item for the vendor across all customers, plus a vendor email.
    vendor_items = defaultdict(list)
    vendor_email_map = {}
    for item in open_items:
        vendor_items[item["vendor"]].append(item)
        if item.get("vendor_email") and not vendor_email_map.get(item["vendor"]):
            vendor_email_map[item["vendor"]] = item["vendor_email"]

    # Emit per-vendor draft payloads (HTML table of all open orders) so the
    # caller can create reviewable Gmail drafts — one per vendor with an email.
    vendor_drafts = []
    for vendor, _count in vendor_po_summary:
        v_email = vendor_email_map.get(vendor, "")
        if not v_email:
            continue
        v_items = vendor_items.get(vendor, [])
        vendor_drafts.append({
            "vendor": vendor,
            "to": v_email,
            "subject": "JIT4You open orders — ETA update request",
            "html_body": build_vendor_email_html(vendor, v_items),
            "item_count": len(v_items),
        })
    try:
        drafts_path = os.path.join(output_dir, "vendor_drafts.json")
        with open(drafts_path, "w") as f:
            json.dump(vendor_drafts, f, indent=2)
        log(f"Vendor drafts payload saved: {drafts_path} ({len(vendor_drafts)} vendors)")
    except Exception as e:
        log(f"  Could not write vendor_drafts.json: {e}")

    # Self-contained sender: the report embeds each vendor's email + a hidden
    # Per-vendor styled HTML emails are sent as separate emails (see
    # send_vendor_drafts); the main report stays clean — no in-report buttons.
    inpage_sender = ""

    # Pre-render the vendor summary rows.
    vendor_summary_rows = ""
    for i, (vendor, count) in enumerate(vendor_po_summary):
        bgv = "#f8f9fa" if i % 2 == 0 else "#ffffff"
        vendor_summary_rows += (
            f'<tr style="background:{bgv};border-bottom:1px solid #e9ecef;">'
            f'<td style="padding:6px 14px;font-size:12px;font-family:Arial,sans-serif;">{vendor}</td>'
            f'<td style="padding:6px 14px;text-align:center;font-size:12px;font-weight:600;'
            f'font-family:Arial,sans-serif;color:#1F4E79;">{count}</td>'
            f'</tr>\n'
        )

    # ---- BUILD HTML ----
    NCOLS = "9"  # SO#, Status, Date, Product, Ord, Del, Open, Pending PO, ETA
    F = "font-size:12px;font-family:Arial,sans-serif;"
    html_rows = ""

    # Track which POs have already had a button rendered, so each PO# gets
    # at most one "Email vendor" button across the whole table (first occurrence).
    pos_with_button = set()
    bcc_for_mailto = CONFIG.get("vendor_followup_bcc", "")

    for customer in sorted(by_customer.keys(), key=str.lower):
        vendors = by_customer[customer]
        all_items = []
        for v in sorted(vendors.keys()):
            all_items.extend(vendors[v])
        cust_sos = len(set(i["so_num"] for i in all_items))

        # CUSTOMER HEADER ROW
        html_rows += (
            f'<tr style="background:#1F4E79;">'
            f'<td colspan="{NCOLS}" style="padding:10px 14px;color:#FFFFFF;font-weight:700;font-size:13px;">'
            f'{customer}'
            f'<span style="font-weight:400;font-size:11px;color:#b0c4de;margin-left:12px;">'
            f'{cust_sos} SO(s) &middot; {len(all_items)} items'
            f'</span></td></tr>\n'
        )

        for vendor in sorted(vendors.keys()):
            items = vendors[vendor]

            # VENDOR SUB-HEADER ROW
            html_rows += (
                f'<tr style="background:#e8eef4;">'
                f'<td colspan="{NCOLS}" style="padding:7px 14px 7px 28px;font-weight:600;'
                f'color:#2c3e50;font-size:12px;border-left:4px solid #1F4E79;">'
                f'{vendor}'
                f'<span style="font-weight:400;font-size:11px;color:#888;margin-left:10px;">'
                f'{len(items)} items'
                f'</span></td></tr>\n'
            )

            # ITEM ROWS
            for idx, item in enumerate(items):
                bg = "#f8f9fa" if idx % 2 == 0 else "#ffffff"
                link = f"{VTIGER_BASE}/view/detail?id={item['so_id']}&module=SalesOrder"

                st = item["so_status"]
                if "Partial" in st:
                    bb, bc = "#fff3cd", "#856404"
                elif st == "Approved":
                    bb, bc = "#d4edda", "#155724"
                else:
                    bb, bc = "#cce5ff", "#004085"

                # Render PO cell. Each PO# gets at most ONE "Email vendor" button across
                # the whole table — rendered the first time the PO# is encountered.
                if item["pending_pos"]:
                    _po_parts = []
                    for _pn in [p.strip() for p in item["pending_pos"].split(",") if p.strip()]:
                        _info = po_email_data.get(_pn, {})
                        if _pn in pos_with_button:
                            _btn = ""  # already shown above
                        else:
                            pos_with_button.add(_pn)
                            _email_url = build_po_email_url(_pn, _info, bcc_for_mailto)
                            if _email_url:
                                _btn = (
                                    f'<a href="{_email_url}" class="po-email-btn" '
                                    f'title="Open an Outlook draft to {_info.get("vendor", "vendor")}" '
                                    f'style="margin-left:6px;padding:1px 8px;font-size:10px;line-height:14px;'
                                    f'border:1px solid #1F4E79;background:#1F4E79;color:#fff !important;'
                                    f'border-radius:3px;cursor:pointer;font-family:Arial;vertical-align:middle;'
                                    f'text-decoration:none;display:inline-block;">'
                                    f'Email vendor</a>'
                                )
                            else:
                                _btn = (
                                    '<span style="margin-left:6px;font-size:10px;color:#999;'
                                    'vertical-align:middle;">(no vendor email)</span>'
                                )
                        _po_parts.append(
                            f'<span style="white-space:nowrap;display:inline-block;margin:1px 0;">'
                            f'<span style="color:#e67e22;">&#9679; {_pn}</span>{_btn}</span>'
                        )
                    po_cell_inner = "<br>".join(_po_parts)
                else:
                    po_cell_inner = '<span style="color:#999;">None</span>'

                # ETA column
                eta_raw = item.get("eta", "")
                if eta_raw:
                    # Colour-code: past = red, within 7 days = orange, future = green
                    try:
                        eta_dt = datetime.strptime(eta_raw.split(" ")[0], "%Y-%m-%d")
                        days_left = (eta_dt - datetime.now()).days
                        eta_text = eta_dt.strftime("%b %d, %Y")
                        eta_color = "#c0392b" if days_left < 0 else ("#e67e22" if days_left <= 7 else "#27ae60")
                    except Exception:
                        eta_text = eta_raw
                        eta_color = "#2c3e50"
                else:
                    eta_text = "—"
                    eta_color = "#999"

                html_rows += (
                    f'<tr style="background:{bg};border-bottom:1px solid #e9ecef;">'
                    f'<td style="padding:6px 8px 6px 28px;{F}font-weight:600;">'
                    f'<a href="{link}" style="color:#1F4E79;text-decoration:none;">{item["so_num"]}</a></td>'
                    f'<td style="padding:6px 8px;{F}">'
                    f'<span style="background:{bb};color:{bc};padding:2px 8px;border-radius:10px;font-size:11px;">{st}</span></td>'
                    f'<td style="padding:6px 8px;{F}">{item["order_date"]}</td>'
                    f'<td style="padding:6px 8px;{F}">{item["product"]}</td>'
                    f'<td style="padding:6px 8px;text-align:center;{F}">{item["ordered_qty"]:g}</td>'
                    f'<td style="padding:6px 8px;text-align:center;{F}">{item["delivered_qty"]:g}</td>'
                    f'<td style="padding:6px 8px;text-align:center;font-weight:600;color:#c0392b;{F}">{item["open_qty"]:g}</td>'
                    f'<td style="padding:6px 8px;{F}">{po_cell_inner}</td>'
                    f'<td style="padding:6px 8px;text-align:center;font-weight:600;{F}color:{eta_color};">{eta_text}</td>'
                    f'</tr>\n'
                )

    email_html = f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>JIT4You Open Orders Report</title></head>
<body style="margin:0;padding:0;font-family:Arial,sans-serif;background:#f0f2f5;">
<table width="100%" cellpadding="0" cellspacing="0" style="max-width:1200px;margin:0 auto;">
<tr><td style="background:#0D2B45;padding:24px 30px;">
<table width="100%" cellpadding="0" cellspacing="0"><tr>
<td><span style="font-size:24px;font-weight:700;color:#FFFFFF;letter-spacing:1px;">JIT4You</span><br>
<span style="font-size:17px;color:#FFFFFF;font-weight:600;">Open Orders by Customer</span></td>
<td style="text-align:right;color:#FFFFFF;font-size:13px;font-weight:500;">{report_date}<br>2026 Sales Orders | Excl. ConMed Customers</td>
</tr></table></td></tr>

<tr><td style="background:#fff;padding:20px 30px;">
<table width="100%" cellpadding="0" cellspacing="8"><tr>
<td style="width:33%;text-align:center;padding:14px;background:#f0f4f8;border-radius:6px;border:1px solid #d0dbe6;">
<div style="font-size:30px;font-weight:700;color:#1F4E79;">{total_customers}</div>
<div style="font-size:12px;color:#555;margin-top:4px;font-weight:600;">Customers</div></td>
<td style="width:33%;text-align:center;padding:14px;background:#f0f4f8;border-radius:6px;border:1px solid #d0dbe6;">
<div style="font-size:30px;font-weight:700;color:#1F4E79;">{total_sos}</div>
<div style="font-size:12px;color:#555;margin-top:4px;font-weight:600;">Open SOs</div></td>
<td style="width:33%;text-align:center;padding:14px;background:#f0f4f8;border-radius:6px;border:1px solid #d0dbe6;">
<div style="font-size:30px;font-weight:700;color:#1F4E79;">{total_items}</div>
<div style="font-size:12px;color:#555;margin-top:4px;font-weight:600;">Open Items</div></td>
</tr></table></td></tr>

<tr><td style="background:#fff;padding:10px 30px 6px;">
<div style="font-size:13px;font-weight:700;color:#1F4E79;margin-bottom:8px;">Open Purchase Orders by Vendor</div>
<table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #d0dbe6;border-collapse:collapse;">
<tr style="background:#0D2B45;">
<th style="padding:8px 14px;color:#FFF;text-align:left;font-size:12px;font-weight:700;">Vendor</th>
<th style="padding:8px 14px;color:#FFF;text-align:center;font-size:12px;font-weight:700;">Open POs</th>
</tr>
{vendor_summary_rows}
</table>
</td></tr>

<tr><td style="background:#fff;padding:10px 30px 20px;">
<table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #dee2e6;border-collapse:collapse;">
<tr style="background:#0D2B45;">
<th style="padding:10px 8px 10px 28px;color:#FFF;text-align:left;font-size:12px;font-weight:700;">SO #</th>
<th style="padding:10px 8px;color:#FFF;text-align:left;font-size:12px;font-weight:700;">Status</th>
<th style="padding:10px 8px;color:#FFF;text-align:left;font-size:12px;font-weight:700;">Date</th>
<th style="padding:10px 8px;color:#FFF;text-align:left;font-size:12px;font-weight:700;">Product</th>
<th style="padding:10px 8px;color:#FFF;text-align:center;font-size:12px;font-weight:700;">Ord</th>
<th style="padding:10px 8px;color:#FFF;text-align:center;font-size:12px;font-weight:700;">Del</th>
<th style="padding:10px 8px;color:#FFF;text-align:center;font-size:12px;font-weight:700;">Open</th>
<th style="padding:10px 8px;color:#FFF;text-align:left;font-size:12px;font-weight:700;">Pending PO</th>
<th style="padding:10px 8px;color:#FFF;text-align:center;font-size:12px;font-weight:700;">ETA</th>
</tr>
{html_rows}
</table></td></tr>

<tr><td style="background:#f0f4f8;padding:16px 30px;text-align:center;">
<span style="font-size:11px;color:#666;">JIT4You Inc. | Open Orders Report | Generated automatically</span>
</td></tr>
</table>
{inpage_sender}
</body></html>"""

    # Save HTML
    html_path = os.path.join(output_dir, "open_orders_report.html")
    with open(html_path, "w") as f:
        f.write(email_html)
    log(f"HTML saved: {html_path}")

    # ---- BUILD EXCEL ----
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="1F4E79")
        dfont = Font(name="Arial", size=10)
        cfmt = '$#,##0.00'
        border = Border(
            left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9")
        )

        ws["A1"] = "JIT4You - Open Orders Report"
        ws["A1"].font = Font(name="Arial", bold=True, size=16, color="1F4E79")
        ws["A2"] = f"Generated: {report_date}"
        ws["A2"].font = Font(name="Arial", size=10, color="666666")

        kpis = [("Customers", total_customers), ("Open SOs", total_sos),
                ("Open Items", total_items)]
        for i, (label, val) in enumerate(kpis):
            rn = 4 + i
            ws[f"A{rn}"] = label
            ws[f"A{rn}"].font = Font(name="Arial", bold=True, size=11)
            c = ws[f"B{rn}"]
            c.value = val
            c.font = Font(name="Arial", size=11, color="1F4E79", bold=True)

        rn = 9
        for col, h in enumerate(["Customer", "Vendor", "Open SOs", "Open Items"], 1):
            c = ws.cell(row=rn, column=col, value=h)
            c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal="center")

        for customer in sorted(by_customer.keys(), key=str.lower):
            for vendor in sorted(by_customer[customer].keys()):
                items = by_customer[customer][vendor]
                rn += 1
                ws.cell(row=rn, column=1, value=customer).font = dfont
                ws.cell(row=rn, column=2, value=vendor).font = dfont
                ws.cell(row=rn, column=3, value=len(set(i["so_num"] for i in items))).font = dfont
                ws.cell(row=rn, column=4, value=len(items)).font = dfont

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12

        # Detail sheet
        ws2 = wb.create_sheet("Open Items Detail")
        headers = ["Customer", "Vendor", "SO Number", "SO Status", "Order Date",
                    "Product", "Ordered Qty", "Delivered Qty", "Open Qty", "Pending PO", "ETA"]
        for col, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal="center")

        alt_fill = PatternFill("solid", fgColor="F2F7FB")
        rn = 2
        for customer in sorted(by_customer.keys(), key=str.lower):
            for vendor in sorted(by_customer[customer].keys()):
                for item in by_customer[customer][vendor]:
                    fill = alt_fill if rn % 2 == 0 else None
                    eta_val = item.get("eta", "") or None
                    vals = [customer, vendor, item["so_num"], item["so_status"], item["order_date"],
                            item["product"], item["ordered_qty"], item["delivered_qty"],
                            item["open_qty"], item["pending_pos"], eta_val]
                    for col, v in enumerate(vals, 1):
                        c = ws2.cell(row=rn, column=col, value=v)
                        c.font = dfont; c.border = border
                        if fill: c.fill = fill
                        if col in (7, 8, 9, 11): c.alignment = Alignment(horizontal="center")
                    rn += 1

        widths = [40, 28, 12, 18, 12, 55, 12, 13, 10, 35, 14]
        for i, w in enumerate(widths):
            ws2.column_dimensions[chr(65 + i)].width = w
        ws2.auto_filter.ref = f"A1:K{rn - 1}"
        ws2.freeze_panes = "A2"

        xlsx_path = os.path.join(output_dir, "open_orders_report.xlsx")
        wb.save(xlsx_path)
        log(f"Excel saved: {xlsx_path}")
    except ImportError:
        log("Warning: openpyxl not installed, skipping Excel generation")
        log("  Install with: pip3 install openpyxl")
        xlsx_path = None

    return email_html, html_path, xlsx_path


# ─────────────────────────────────────────────
# EMAIL VIA ZAPIER
# ─────────────────────────────────────────────
def send_email(email_html, xlsx_path):
    """Send report via Zapier webhook (same format as vtiger_qb_sync.py)."""
    webhook_url = CONFIG.get("zapier_email_webhook", "")
    if not webhook_url:
        log("No Zapier webhook configured — skipping email send")
        return

    subject = f"JIT4You Open Orders Report — {datetime.now().strftime('%A, %B %d, %Y')}"
    payload = {
        "subject": subject,
        "html_body": email_html,
        "to": "customersupport@jit4you.com",
    }

    try:
        http_request(webhook_url, method="POST", json_body=payload)
        log("Email sent via Zapier webhook")
    except Exception as e:
        log(f"Failed to send email: {e}")


def send_vendor_drafts(output_dir):
    """Send one HTML-table email per vendor (all of that vendor's open orders)
    via the Zapier webhook to the internal review inbox (CONFIG['vendor_drafts_to']).
    These are addressed to the user, NOT the vendors — for manual review/forwarding.
    Reads the vendor_drafts.json payload written by generate_report().
    """
    to_addr = CONFIG.get("vendor_drafts_to", "")
    webhook_url = CONFIG.get("zapier_email_webhook", "")
    if not to_addr or not webhook_url:
        log("Vendor drafts: no recipient/webhook configured — skipping")
        return

    path = os.path.join(output_dir, "vendor_drafts.json")
    if not os.path.exists(path):
        log("Vendor drafts: vendor_drafts.json not found — skipping")
        return
    try:
        with open(path) as f:
            drafts = json.load(f)
    except Exception as e:
        log(f"Vendor drafts: could not load payload: {e}")
        return

    sent = 0
    for d in drafts:
        payload = {
            "subject": f"[Open-orders draft → {d['vendor']} | {d['to']}] ETA update request",
            "html_body": d["html_body"],
            "to": to_addr,
        }
        try:
            http_request(webhook_url, method="POST", json_body=payload)
            sent += 1
            log(f"  Vendor draft sent: {d['vendor']} ({d['item_count']} items) → {to_addr}")
        except Exception as e:
            log(f"  Failed to send vendor draft for {d['vendor']}: {e}")
    log(f"Vendor draft emails sent: {sent}/{len(drafts)} → {to_addr}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="JIT4You Open Orders Report")
    parser.add_argument("--no-email", action="store_true", help="Skip sending email")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, no report generation")
    args = parser.parse_args()

    log("=" * 60)
    log("JIT4You Open Orders by Customer Report")
    log("=" * 60)

    # Connect to Vtiger REST API. Cache file is date-scoped so each day's
    # scheduled run starts fresh while intra-day resume runs share progress.
    today = datetime.now().strftime("%Y-%m-%d")
    cache_path = os.path.join(CONFIG["output_dir"], f"retrieve_cache_{today}.json")
    # Clean up stale cache files from previous days
    try:
        import glob as _glob
        for old in _glob.glob(os.path.join(CONFIG["output_dir"], "retrieve_cache_*.json")):
            if old != cache_path:
                os.remove(old)
    except Exception:
        pass
    vt = VtigerAPI(CONFIG["vtiger_rest_base"], CONFIG["vtiger_user"],
                   CONFIG["vtiger_accesskey"], cache_path=cache_path)
    vt.login()

    # Extract data
    open_items = extract_open_orders(vt, dry_run=args.dry_run)

    if args.dry_run:
        log("Dry run complete")
        return

    # Completeness gate: Vtiger rate-limits hard and each run is time-capped, so
    # a single pass may not fetch every record. The persistent cache lets repeated
    # runs make progress. Only generate/send the report when this run fetched
    # everything cleanly (no failures) — otherwise exit so the next run resumes.
    if vt.fetch_failures > 0:
        vt.save_cache()
        log(f"INCOMPLETE: {vt.fetch_failures} record fetches failed (likely rate-limited).")
        log("Progress saved to cache. Re-run to resume; report not generated/sent this pass.")
        sys.exit(2)

    if not open_items:
        log("No open items found!")
        return

    log(f"\n{'=' * 60}")
    log(f"RESULTS: {len(open_items)} open items, "
        f"{len(set(i['customer'] for i in open_items))} customers")
    log(f"{'=' * 60}\n")

    # Generate report
    email_html, html_path, xlsx_path = generate_report(open_items, CONFIG["output_dir"])

    # Send the main report, then one separate styled HTML email per vendor.
    if not args.no_email:
        send_email(email_html, xlsx_path)
        send_vendor_drafts(CONFIG["output_dir"])
    else:
        log("Skipping email (--no-email flag)")

    log("Done!")


if __name__ == "__main__":
    main()
